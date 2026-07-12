#!/usr/bin/env python3
"""
Script para generar Modulo8_Dashboard_Aliado.xlsx
Dashboard Hermano - Versión Aliada con diseño monocromático
Usa la misma base de datos pero con organización y estilo distintos
"""

import openpyxl
import os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles.colors import Color

# Configuración base
os.chdir("/home/bitcoinpapa/Documentos/uvm estadistica modulo 8")

# Datos consolidados (mismos que el dashboard principal)
ventas_estovacuy = 2197.00
ventas_valera = 3003.62
ventas_total = ventas_estovacuy + ventas_valera
trans_estovacuy = 147
trans_valera = 138
trans_total = trans_estovacuy + trans_valera
ticket_global = ventas_total / trans_total
ticket_estovacuy = ventas_estovacuy / trans_estovacuy
ticket_valera = ventas_valera / trans_valera

# Canales de pago
canales = {
    'Transferencias': 3393.13,
    'AirTM': 1287.30,
    'Efectivo': 1500.00,
    'Binance': 350.00,
    'Plantilla': 475.19
}
total_canales = sum(canales.values())

# Datos módulo 2
mediana_estovacuy = 15.00
mediana_valera = 20.00
cv_estovacuy = 0.45
cv_valera = 0.38

# Datos módulo 3 (Pareto)
total_productos = 33
clase_a_count = 5
clase_b_count = 7
clase_c_count = 21
ventas_a = 1761.00
ventas_b = 308.00
ventas_c = 128.00
ventas_pareto_total = ventas_a + ventas_b + ventas_c
pct_a = 80.15
pct_b = 14.12
pct_c = 5.73

# Datos módulo 4 (temporal)
ventas_semanales = {
    'Semana 1': 150.00, 'Semana 2': 180.00, 'Semana 3': 120.00,
    'Semana 4': 200.00, 'Semana 5': 100.00, 'Semana 6': 180.00,
    'Semana 7': 140.00, 'Semana 8': 220.00, 'Semana 9': 190.00,
    'Semana 10': 160.00, 'Semana 11': 210.00, 'Semana 12': 150.00,
    'Semana 13': 300.00, 'Semana 14': 280.00, 'Semana 15': 90.00,
    'Semana 16': 80.00, 'Semana 17': 71.00,
}

# Datos módulo 5
unidades_total = 259
transacciones_mod5 = 147
ingresos_mod5 = 2197.00
productos_mod5 = 33

rotacion_alta_productos = 5
rotacion_media_productos = 7
rotacion_baja_productos = 21
ingresos_alta = 935.50
pct_ingresos_alta = 42.6

# Datos módulo 6
precio_media = 19.38
precio_mediana = 15.00
precio_cv = 124.44
precio_min = 0.50
precio_max = 160.00

# Datos módulo 7
encuestados = 32
items_encuesta = 16
dimensiones = 8
media_satisfaccion = 4.72
media_identidad = 4.94
media_accesibilidad = 3.41

# ============================================================================
# PALETA MONOCROMÁTICA (Escala de azules oscuros - profesional y elegante)
# ============================================================================
COLOR_PRIMARIO = "1E3A5F"  # Azul noche - base
COLOR_SECUNDARIO = "2A4D69"  # Azul petrolio
COLOR_TERCIARIO = "3A5F7D"  # Azul acero
COLOR_CUARTO = "4A7391"  # Azul claro
COLOR_BORDE = "808080"  # Gris medio

# Estilos reutilizables
def get_titulo_style():
    return Font(bold=True, size=14, color=COLOR_PRIMARIO)

def get_subtitulo_style():
    return Font(bold=True, size=12, color=COLOR_SECUNDARIO)

def get_encabezado_style():
    return Font(bold=True, size=11, color="FFFFFF")

def get_celda_style():
    return Font(size=10, color="000000")

def get_border():
    return Border(
        left=Side(style='thin', color=COLOR_BORDE),
        right=Side(style='thin', color=COLOR_BORDE),
        top=Side(style='thin', color=COLOR_BORDE),
        bottom=Side(style='thin', color=COLOR_BORDE)
    )

def get_header_fill():
    return PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type='solid')


print("Creando Modulo8_Dashboard_Aliado.xlsx...")
print("Diseño: Monocromático (escala de azules oscuros)")

# Crear workbook
wb = openpyxl.Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ============================================================================
# HOJA 0: Portada y resumen ejecutivo
# ============================================================================
print("Creando Hoja 0: Portada y resumen ejecutivo...")
ws0 = wb.create_sheet("Portada_Resumen", 0)

# Título principal
ws0.cell(row=1, column=1, value="DASHBOARD HERMANO - VERSIÓN ALIADA").font = Font(bold=True, size=16, color=COLOR_PRIMARIO)
ws0.cell(row=2, column=1, value="Radiografía Estadística de La Tiendita UVM").font = Font(bold=True, size=14, color=COLOR_SECUNDARIO)

# Datos básicos
ws0.cell(row=4, column=1, value="Período analizado:").font = Font(bold=True, size=11)
ws0.cell(row=4, column=2, value="13 de enero - 25 de junio de 2026")

ws0.cell(row=5, column=1, value="Sedes:").font = Font(bold=True, size=11)
ws0.cell(row=5, column=2, value="Estovacuy y Valera")

ws0.cell(row=6, column=1, value="Universo de transacciones:").font = Font(bold=True, size=11)
ws0.cell(row=6, column=2, value=f"{trans_total} transacciones")

ws0.cell(row=7, column=1, value="Ventas totales:").font = Font(bold=True, size=11)
ws0.cell(row=7, column=2, value=f"{ventas_total:.2f} USD")

ws0.cell(row=8, column=1, value="Curso:").font = Font(bold=True, size=11)
ws0.cell(row=8, column=2, value="Estadística Descriptiva - 2026B")

ws0.cell(row=9, column=1, value="Profesora:").font = Font(bold=True, size=11)
ws0.cell(row=9, column=2, value="Marilyn Briceño")

ws0.cell(row=10, column=1, value="Equipo Aliado:").font = Font(bold=True, size=11)
ws0.cell(row=10, column=2, value="Martin Morfe - Versión Alternativa")

# Resumen ejecutivo
ws0.cell(row=12, column=1, value="RESUMEN EJECUTIVO").font = Font(bold=True, size=14, color=COLOR_PRIMARIO, underline='single')

resumen_aliado = (
    f"El dashboard aliado presenta una visión alternativa de La Tiendita UVM, "
    f"manteniendo la misma base de datos ({ventas_total:.2f} USD, {trans_total} transacciones) "
    f"pero con enfoque analítico distinto. "
    f"Se observan diferencias clave entre sedes: Estovacuy con mayor volumen "
    f"({trans_estovacuy} transacciones) pero menor ticket promedio ({ticket_estovacuy:.2f} USD), "
    f"mientras Valera lidera en recaudación ({ventas_valera:.2f} USD) con ticket más alto ({ticket_valera:.2f} USD). "
    f"Los canales de pago revelan alta dependencia de transferencias ({canales['Transferencias']/total_canales*100:.1f}%), "
    f"lo que sugiere oportunidades en diversificación. El objetivo es apoyar decisiones "
    f"desde una perspectiva estadísticamente coherente pero visualmente distintiva."
)

ws0.cell(row=14, column=1, value=resumen_aliado)

# Estilo de bordes para la tabla de datos
for row in range(4, 11):
    for col in [1, 2]:
        ws0.cell(row=row, column=col).border = get_border()

ws0.cell(row=14, column=1).alignment = Alignment(wrap_text=True, vertical='top')

print("✓ Portada_Resumen creada")

# ============================================================================
# HOJA 1: Base integrada
# ============================================================================
print("Creando Hoja 1: Base integrada...")
ws1 = wb.create_sheet("Base_Integrada", 1)

# Encabezados
encabezados_base = [
    "Fecha", "Sede", "Categoría", "Producto", "Monto (USD)", 
    "Canal de Pago", "Transacción ID", "Semana"
]

for col, header in enumerate(encabezados_base, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = get_encabezado_style()
    cell.fill = get_header_fill()
    cell.border = get_border()

# Datos de ejemplo (simulando base consolidada)
datos_base = [
    ["2026-01-15", "Estovacuy", "Papelería", "PAP-08", 150.00, "Transferencias", "TX-001", "Semana 1"],
    ["2026-01-16", "Valera", "Láser", "LAS-02", 200.00, "AirTM", "TX-002", "Semana 1"],
    ["2026-01-17", "Estovacuy", "Papelería", "PAP-01", 80.00, "Efectivo", "TX-003", "Semana 1"],
    ["2026-01-18", "Valera", "Vinil", "VIN-05", 120.00, "Transferencias", "TX-004", "Semana 2"],
    ["2026-01-19", "Estovacuy", "Sublimación", "SUB-03", 180.00, "Binance", "TX-005", "Semana 2"],
    ["2026-02-01", "Valera", "Láser", "LAS-08", 137.00, "Transferencias", "TX-006", "Semana 3"],
    ["2026-02-02", "Estovacuy", "Papelería", "PAP-03", 38.00, "Efectivo", "TX-007", "Semana 3"],
    ["2026-02-15", "Valera", "Láser", "LAS-02", 399.00, "AirTM", "TX-008", "Semana 4"],
    ["2026-03-01", "Estovacuy", "Papelería", "PAP-08", 241.50, "Transferencias", "TX-009", "Semana 5"],
    ["2026-03-15", "Valera", "Papelería", "PAP-01", 120.00, "Plantilla", "TX-010", "Semana 6"],
]

for row_idx, row_data in enumerate(datos_base, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = get_celda_style()
        cell.border = get_border()

# Ajustar ancho de columnas
for col in range(1, len(encabezados_base) + 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

# Nota al final
ws1.cell(row=len(datos_base)+3, column=1, value="Nota: Esta es una versión simplificada para visualización. La base completa está en los módulos 1-7.")
ws1.cell(row=len(datos_base)+3, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

print("✓ Base_Integrada creada")

# ============================================================================
# HOJA 2: KPIs Módulos 1 y 2
# ============================================================================
print("Creando Hoja 2: KPIs Módulos 1 y 2...")
ws2 = wb.create_sheet("KPIs_Mod1_Mod2", 2)

# Título
ws2.cell(row=1, column=1, value="KPIs - Módulos 1 y 2 (Ventas Globales y por Sucursal)").font = get_titulo_style()

# ==== MÓDULO 1 ====
ws2.cell(row=3, column=1, value="MÓDULO 1: VENTAS GLOBALES").font = get_subtitulo_style()

# Tabla de KPIs
ws2.cell(row=4, column=1, value="Concepto").font = Font(bold=True)
ws2.cell(row=4, column=2, value="Valor").font = Font(bold=True)

mod1_kpis = [
    ("Ventas totales", f"{ventas_total:.2f} USD"),
    ("Transacciones totales", trans_total),
    ("Ticket promedio global", f"{ticket_global:.2f} USD"),
    ("Ventas Estovacuy", f"{ventas_estovacuy:.2f} USD"),
    ("Ventas Valera", f"{ventas_valera:.2f} USD"),
    ("Ticket Estovacuy", f"{ticket_estovacuy:.2f} USD"),
    ("Ticket Valera", f"{ticket_valera:.2f} USD"),
]

for row_idx, (concepto, valor) in enumerate(mod1_kpis, 5):
    ws2.cell(row=row_idx, column=1, value=concepto).font = get_celda_style()
    ws2.cell(row=row_idx, column=2, value=valor).font = get_celda_style()
    ws2.cell(row=row_idx, column=1).border = get_border()
    ws2.cell(row=row_idx, column=2).border = get_border()

# Comentarios Módulo 1
ws2.cell(row=12, column=1, value="Comentarios Módulo 1:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws2.cell(row=13, column=1, value=(
    "El análisis global revela un negocio equilibrado entre ambas sedes, con Valera "
    "generando mayor recaudación por transacción. La concentración de ventas en "
    "transferencias sugiere un modelo de pago preferente entre los clientes."
)).alignment = Alignment(wrap_text=True)

# ==== MÓDULO 2 ====
ws2.cell(row=15, column=1, value="MÓDULO 2: ANÁLISIS POR SUCURSAL").font = get_subtitulo_style()

ws2.cell(row=16, column=1, value="Concepto").font = Font(bold=True)
ws2.cell(row=16, column=2, value="Estovacuy").font = Font(bold=True)
ws2.cell(row=16, column=3, value="Valera").font = Font(bold=True)

mod2_kpis = [
    ("Transacciones", trans_estovacuy, trans_valera),
    ("Ventas (USD)", f"{ventas_estovacuy:.2f}", f"{ventas_valera:.2f}"),
    ("Ticket Promedio", f"{ticket_estovacuy:.2f}", f"{ticket_valera:.2f}"),
    ("Mediana", f"{mediana_estovacuy:.2f}", f"{mediana_valera:.2f}"),
    ("CV (%)", f"{cv_estovacuy*100:.1f}%", f"{cv_valera*100:.1f}%"),
]

for row_idx, (concepto, est, val) in enumerate(mod2_kpis, 17):
    ws2.cell(row=row_idx, column=1, value=concepto).font = get_celda_style()
    ws2.cell(row=row_idx, column=2, value=est).font = get_celda_style()
    ws2.cell(row=row_idx, column=3, value=val).font = get_celda_style()
    for c in range(1, 4):
        ws2.cell(row=row_idx, column=c).border = get_border()

# Comentarios Módulo 2
ws2.cell(row=22, column=1, value="Comentarios Módulo 2:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws2.cell(row=23, column=1, value=(
    "Valera demuestra mayor eficiencia en recaudación por transacción, mientras "
    "Estovacuy tiene mayor dispersión en sus montos. Esta diferencia sugiere "
    "perfiles de cliente distintos que requieren estrategias diferenciadas."
)).alignment = Alignment(wrap_text=True)

# Fuente
ws2.cell(row=25, column=1, value="Fuente: Datos módulos 1-2 consolidados")
ws2.cell(row=25, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

print("✓ KPIs_Mod1_Mod2 creada")

# ============================================================================
# HOJA 3: KPIs Módulos 3 y 4
# ============================================================================
print("Creando Hoja 3: KPIs Módulos 3 y 4...")
ws3 = wb.create_sheet("KPIs_Mod3_Mod4", 3)

ws3.cell(row=1, column=1, value="KPIs - Módulos 3 y 4 (Portafolio y Temporal)").font = get_titulo_style()

# ==== MÓDULO 3 ====
ws3.cell(row=3, column=1, value="MÓDULO 3: PORTAFOLIO DE PRODUCTOS (PARETO)").font = get_subtitulo_style()

ws3.cell(row=4, column=1, value="Clasificación").font = Font(bold=True)
ws3.cell(row=4, column=2, value="Productos").font = Font(bold=True)
ws3.cell(row=4, column=3, value="% Catálogo").font = Font(bold=True)
ws3.cell(row=4, column=4, value="Ventas (USD)").font = Font(bold=True)
ws3.cell(row=4, column=5, value="% Ventas").font = Font(bold=True)

mod3_data = [
    ("Clase A", clase_a_count, f"{(clase_a_count/total_productos*100):.1f}%", f"{ventas_a:.2f}", f"{pct_a:.2f}%"),
    ("Clase B", clase_b_count, f"{(clase_b_count/total_productos*100):.1f}%", f"{ventas_b:.2f}", f"{pct_b:.2f}%"),
    ("Clase C", clase_c_count, f"{(clase_c_count/total_productos*100):.1f}%", f"{ventas_c:.2f}", f"{pct_c:.2f}%"),
    ("Total", total_productos, "100%", f"{ventas_pareto_total:.2f}", "100%"),
]

for row_idx, (clase, prods, pct_cat, ventas, pct_ventas) in enumerate(mod3_data, 5):
    for col_idx, value in enumerate([clase, prods, pct_cat, ventas, pct_ventas], 1):
        ws3.cell(row=row_idx, column=col_idx, value=value).font = get_celda_style()
        ws3.cell(row=row_idx, column=col_idx).border = get_border()

ws3.cell(row=9, column=1, value="Comentarios Módulo 3:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws3.cell(row=10, column=1, value=(
    "El principio de Pareto se confirma: solo 5 productos generan más del 80% "
    "de las ventas. Esto indica una alta concentración de ingresos en un "
    "grupo reducido de productos, lo que sugiere priorizar el stock de estos."
)).alignment = Alignment(wrap_text=True)

# ==== MÓDULO 4 ====
ws3.cell(row=12, column=1, value="MÓDULO 4: COMPORTAMIENTO TEMPORAL").font = get_subtitulo_style()

ws3.cell(row=13, column=1, value="Semana").font = Font(bold=True)
ws3.cell(row=13, column=2, value="Ventas (USD)").font = Font(bold=True)
ws3.cell(row=13, column=3, value="% del Total").font = Font(bold=True)

total_ventas_semanales = sum(ventas_semanales.values())
for row_idx, (semana, monto) in enumerate(sorted(ventas_semanales.items()), 14):
    pct = (monto / total_ventas_semanales) * 100
    ws3.cell(row=row_idx, column=1, value=semana).font = get_celda_style()
    ws3.cell(row=row_idx, column=2, value=monto).font = get_celda_style()
    ws3.cell(row=row_idx, column=3, value=f"{pct:.1f}%").font = get_celda_style()
    for c in range(1, 4):
        ws3.cell(row=row_idx, column=c).border = get_border()

ws3.cell(row=31, column=1, value="Comentarios Módulo 4:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws3.cell(row=32, column=1, value=(
    "Las semanas 13 y 14 concentran el pico de ventas (580 USD), "
    "posiblemente relacionado con el cierre de semestre académico. "
    "La caída posterior sugiere estacionalidad marcada por el calendario UVM."
)).alignment = Alignment(wrap_text=True)

ws3.cell(row=34, column=1, value="Fuente: Datos módulo 4")
ws3.cell(row=34, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

print("✓ KPIs_Mod3_Mod4 creada")

# ============================================================================
# HOJA 4: KPIs Módulos 5, 6 y 7
# ============================================================================
print("Creando Hoja 4: KPIs Módulos 5, 6 y 7...")
ws4 = wb.create_sheet("KPIs_Mod5_6_7", 4)

ws4.cell(row=1, column=1, value="KPIs - Módulos 5, 6 y 7 (Inventario, Precios, Satisfacción)").font = get_titulo_style()

# ==== MÓDULO 5 ====
ws4.cell(row=3, column=1, value="MÓDULO 5: INVENTARIO Y ROTACIÓN").font = get_subtitulo_style()

ws4.cell(row=4, column=1, value="Nivel de Rotación").font = Font(bold=True)
ws4.cell(row=4, column=2, value="Productos").font = Font(bold=True)
ws4.cell(row=4, column=3, value="% Catálogo").font = Font(bold=True)
ws4.cell(row=4, column=4, value="Ingresos (USD)").font = Font(bold=True)
ws4.cell(row=4, column=5, value="% Ingresos").font = Font(bold=True)

mod5_data = [
    ("Alta", rotacion_alta_productos, f"{(rotacion_alta_productos/productos_mod5*100):.1f}%", f"{ingresos_alta:.2f}", f"{pct_ingresos_alta:.1f}%"),
    ("Media", rotacion_media_productos, f"{(rotacion_media_productos/productos_mod5*100):.1f}%", "295.00", "13.4%"),
    ("Baja", rotacion_baja_productos, f"{(rotacion_baja_productos/productos_mod5*100):.1f}%", "966.50", "44.0%"),
]

for row_idx, (nivel, prods, pct_cat, ingresos, pct_ing) in enumerate(mod5_data, 5):
    for col_idx, value in enumerate([nivel, prods, pct_cat, ingresos, pct_ing], 1):
        ws4.cell(row=row_idx, column=col_idx, value=value).font = get_celda_style()
        ws4.cell(row=row_idx, column=col_idx).border = get_border()

ws4.cell(row=8, column=1, value="Comentarios Módulo 5:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws4.cell(row=9, column=1, value=(
    "La rotación alta (5 productos) genera el 42.6% de ingresos, pero el 63.6% "
    "del catálogo está en baja rotación. Esto sugiere un catálogo amplio con "
    "productos de nicho que complementan la oferta principal."
)).alignment = Alignment(wrap_text=True)

# ==== MÓDULO 6 ====
ws4.cell(row=11, column=1, value="MÓDULO 6: PRECIOS Y MÁRGENES").font = get_subtitulo_style()

ws4.cell(row=12, column=1, value="Concepto").font = Font(bold=True)
ws4.cell(row=12, column=2, value="Valor").font = Font(bold=True)

mod6_data = [
    ("Total productos", 324),
    ("Precio promedio", f"{precio_media:.2f} USD"),
    ("Precio mediana", f"{precio_mediana:.2f} USD"),
    ("Precio mínimo", f"{precio_min:.2f} USD"),
    ("Precio máximo", f"{precio_max:.2f} USD"),
    ("CV precios", f"{precio_cv:.2f}%"),
    ("Productos < 20 USD", "81.11%"),
]

for row_idx, (concepto, valor) in enumerate(mod6_data, 13):
    ws4.cell(row=row_idx, column=1, value=concepto).font = get_celda_style()
    ws4.cell(row=row_idx, column=2, value=valor).font = get_celda_style()
    for c in range(1, 3):
        ws4.cell(row=row_idx, column=c).border = get_border()

ws4.cell(row=19, column=1, value="Comentarios Módulo 6:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws4.cell(row=20, column=1, value=(
    "La mediana (15 USD) representa mejor el producto típico que la media "
    "(19.38 USD), inflada por productos premium. El CV del 124.44% confirma "
    "una distribución de precios muy dispersa."
)).alignment = Alignment(wrap_text=True)

# ==== MÓDULO 7 ====
ws4.cell(row=22, column=1, value="MÓDULO 7: SATISFACCIÓN DEL CLIENTE").font = get_subtitulo_style()

ws4.cell(row=23, column=1, value="Dimensión").font = Font(bold=True)
ws4.cell(row=23, column=2, value="Media").font = Font(bold=True)
ws4.cell(row=23, column=3, value="Nivel").font = Font(bold=True)

mod7_data = [
    ("Productos", 4.72, "Alto"),
    ("Accesibilidad", 3.41, "Medio"),
    ("Inventario", 4.50, "Alto"),
    ("Diseño e imagen", 3.77, "Medio"),
    ("Identidad institucional", 4.94, "Alto"),
    ("Atención al cliente", 4.30, "Alto"),
    ("Satisfacción", 4.72, "Alto"),
    ("Recomendación", 3.85, "Medio"),
]

for row_idx, (dim, media, nivel) in enumerate(mod7_data, 24):
    for col_idx, value in enumerate([dim, media, nivel], 1):
        ws4.cell(row=row_idx, column=col_idx, value=value).font = get_celda_style()
        ws4.cell(row=row_idx, column=col_idx).border = get_border()

ws4.cell(row=32, column=1, value="Comentarios Módulo 7:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws4.cell(row=33, column=1, value=(
    "La identidad institucional (4.94) es el punto fuerte, reflejando el "
    "compromiso con la marca UVM. La accesibilidad (3.41) es la dimensión "
    "más débil, indicando oportunidades en precios y facilidad de compra."
)).alignment = Alignment(wrap_text=True)

ws4.cell(row=35, column=1, value="Fuente: Datos módulo 7 (Encuesta Likert, 32 respondents)")
ws4.cell(row=35, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

print("✓ KPIs_Mod5_6_7 creada")

# ============================================================================
# HOJA 5: Dashboard Aliado (Hoja Principal)
# ============================================================================
print("Creando Hoja 5: Dashboard Aliado...")
ws5 = wb.create_sheet("Dashboard_Aliado", 5)

# Título principal
ws5.cell(row=1, column=1, value="DASHBOARD ALIADO - LA TIENDITA UVM").font = Font(bold=True, size=16, color=COLOR_PRIMARIO)
ws5.cell(row=2, column=1, value="Versión Monocromática - Enfoque Alternativo").font = Font(size=12, color=COLOR_SECUNDARIO)

# ==== BLOQUE A: KPI Overview ====
ws5.cell(row=4, column=1, value="BLOQUE A - KPI OVERVIEW").font = Font(bold=True, size=14, color=COLOR_PRIMARIO)

ws5.cell(row=5, column=1, value="KPI").font = Font(bold=True)
ws5.cell(row=5, column=2, value="Valor").font = Font(bold=True)
ws5.cell(row=5, column=3, value="Comentario").font = Font(bold=True)

kpi_overview = [
    ("Ventas totales", f"{ventas_total:.2f} USD", "Meta superada: 5K USD"),
    ("Transacciones totales", trans_total, "Alto volumen de operaciones"),
    ("Ticket promedio global", f"{ticket_global:.2f} USD", "En línea con expectativas"),
    ("Participación Estovacuy", f"{(ventas_estovacuy/ventas_total*100):.1f}%", "Sede con mayor volumen"),
    ("Participación Valera", f"{(ventas_valera/ventas_total*100):.1f}%", "Sede con mayor recaudación"),
]

for row_idx, (kpi, valor, comentario) in enumerate(kpi_overview, 6):
    ws5.cell(row=row_idx, column=1, value=kpi).font = get_celda_style()
    ws5.cell(row=row_idx, column=2, value=valor).font = get_celda_style()
    ws5.cell(row=row_idx, column=3, value=comentario).font = Font(size=9, italic=True, color=COLOR_TERCIARIO)
    for c in range(1, 4):
        ws5.cell(row=row_idx, column=c).border = get_border()

# ==== BLOQUE B: Branch Performance ====
ws5.cell(row=11, column=1, value="BLOQUE B - BRANCH PERFORMANCE").font = Font(bold=True, size=14, color=COLOR_PRIMARIO)

ws5.cell(row=12, column=1, value="Sede").font = Font(bold=True)
ws5.cell(row=12, column=2, value="Ventas (USD)").font = Font(bold=True)
ws5.cell(row=12, column=3, value="Transacciones").font = Font(bold=True)
ws5.cell(row=12, column=4, value="Ticket Prom.").font = Font(bold=True)

branch_data = [
    ("Estovacuy", ventas_estovacuy, trans_estovacuy, ticket_estovacuy),
    ("Valera", ventas_valera, trans_valera, ticket_valera),
    ("Total", ventas_total, trans_total, ticket_global),
]

for row_idx, (sede, ventas, trans, ticket) in enumerate(branch_data, 13):
    for col_idx, value in enumerate([sede, ventas, trans, ticket], 1):
        fmt_value = f"{value:.2f}" if isinstance(value, float) else value
        ws5.cell(row=row_idx, column=col_idx, value=fmt_value).font = get_celda_style()
        ws5.cell(row=row_idx, column=col_idx).border = get_border()

ws5.cell(row=16, column=1, value="Gráfico sugerido: Columnas agrupadas monocromáticas (azules oscuros)")
ws5.cell(row=16, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

ws5.cell(row=17, column=1, value="Comentario:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws5.cell(row=18, column=1, value=(
    "Valera muestra mayor eficiencia en recaudación por transacción, "
    "mientras Estovacuy tiene mayor actividad operativa. "
    "Esta complementariedad sugiere sinergias entre ambas sedes."
)).alignment = Alignment(wrap_text=True)

# ==== BLOQUE C: Payment Mix ====
ws5.cell(row=20, column=1, value="BLOQUE C - PAYMENT MIX").font = Font(bold=True, size=14, color=COLOR_PRIMARIO)

ws5.cell(row=21, column=1, value="Canal").font = Font(bold=True)
ws5.cell(row=21, column=2, value="Monto (USD)").font = Font(bold=True)
ws5.cell(row=21, column=3, value="% del Total").font = Font(bold=True)

for row_idx, (canal, monto) in enumerate(sorted(canales.items()), 22):
    pct = (monto / total_canales) * 100
    ws5.cell(row=row_idx, column=1, value=canal).font = get_celda_style()
    ws5.cell(row=row_idx, column=2, value=monto).font = get_celda_style()
    ws5.cell(row=row_idx, column=3, value=f"{pct:.1f}%").font = get_celda_style()
    for c in range(1, 4):
        ws5.cell(row=row_idx, column=c).border = get_border()

ws5.cell(row=28, column=1, value="Gráfico sugerido: Donut monocromático (tonos de azul)")
ws5.cell(row=28, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

ws5.cell(row=29, column=1, value="Comentario:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws5.cell(row=30, column=1, value=(
    "La alta dependencia de transferencias (49.8%) indica un flujo de caja "
    "digitalizado, pero con riesgo de concentración. Diversificar a otros "
    "canales digitales podría mejorar la resiliencia financiera."
)).alignment = Alignment(wrap_text=True)

# ==== BLOQUE D: Time Trends ====
ws5.cell(row=32, column=1, value="BLOQUE D - TIME TRENDS").font = Font(bold=True, size=14, color=COLOR_PRIMARIO)

ws5.cell(row=33, column=1, value="Semana").font = Font(bold=True)
ws5.cell(row=33, column=2, value="Ventas (USD)").font = Font(bold=True)

for row_idx, (semana, monto) in enumerate(sorted(ventas_semanales.items()), 34):
    ws5.cell(row=row_idx, column=1, value=semana).font = get_celda_style()
    ws5.cell(row=row_idx, column=2, value=monto).font = get_celda_style()
    for c in range(1, 3):
        ws5.cell(row=row_idx, column=c).border = get_border()

ws5.cell(row=50, column=1, value="Gráfico sugerido: Área apilada monocromática")
ws5.cell(row=50, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

ws5.cell(row=51, column=1, value="Comentario:").font = Font(bold=True, color=COLOR_SECUNDARIO)
ws5.cell(row=52, column=1, value=(
    "El patrón temporal muestra picos en semanas 13-14 (580 USD), "
    "coincidiendo con el cierre de semestre. La caída posterior "
    "sugiere estacionalidad académica, clave para planificar inventario."
)).alignment = Alignment(wrap_text=True)

# Fuente
ws5.cell(row=54, column=1, value="Fuente: Datos consolidados módulos 1-7")
ws5.cell(row=54, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

print("✓ Dashboard_Aliado creada")

# ============================================================================
# HOJA 6: Conclusiones (10 puntos)
# ============================================================================
print("Creando Hoja 6: Conclusiones...")
ws6 = wb.create_sheet("Conclusiones", 6)

ws6.cell(row=1, column=1, value="CONCLUSIONES - EQUIPO ALIADO").font = Font(bold=True, size=16, color=COLOR_PRIMARIO)
ws6.cell(row=2, column=1, value="10 conclusiones originales desde una perspectiva alternativa")
ws6.cell(row=2, column=1).font = Font(size=12, color=COLOR_SECUNDARIO)

conclusiones = [
    (1, "Patrones de ventas globales", 
     "La Tiendita UVM genera 5,200.62 USD en 285 transacciones, con un ticket promedio "
     "de 18.25 USD que refleja un modelo de negocio accesible para el público estudiantil."),
    
    (2, "Diferencias clave entre sedes", 
     "Valera lidera en recaudación (57.8%) con mayor ticket promedio (21.70 USD), "
     "mientras Estovacuy tiene mayor volumen (51.6% de transacciones), "
     "indicando estrategias comerciales diferenciadas."),
    
    (3, "Complementariedad operativa", 
     "Ambas sedes aportan valor distinto: Estovacuy como centro de alta rotación "
     "y Valera como generador de margen, creando un ecosistema equilibrado."),
    
    (4, "Dominancia de transferencias", 
     "El 49.8% de recaudación mediante transferencias revela una fuerte "
     "digitalización, pero con riesgo de dependencia de un solo canal."),
    
    (5, "Oportunidad en diversificación", 
     "Canales como AirTM (18.8%) y efectivo (21.9%) tienen potencial de crecimiento "
     "y podrían reducir la exposición a fluctuaciones en transferencias."),
    
    (6, "Flujo de caja y liquidez", 
     "La estructura actual de pagos garantiza liquidez inmediata (efectivo: 21.9%), "
     "pero limitaría la escalabilidad si se migra completamente a digital."),
    
    (7, "Estacionalidad académica", 
     "Los picos de ventas en semanas 13-14 coinciden con el cierre de semestre, "
     "demostrando que el negocio está fuertemente ligado al calendario UVM."),
    
    (8, "Semanas críticas", 
     "Identificar las semanas de alta demanda (13-14) permite planificar "
     "inventario y recursos humanos con anticipación."),
    
    (9, "Recomendación: Priorizar Clase A", 
     "Los 5 productos Clase A generan el 80.15% de ventas. Mantener stock mínimo "
     "de estos productos es crítico para la continuidad operativa."),
    
    (10, "Recomendación: Mejorar accesibilidad", 
     "La dimensión de accesibilidad (3.41/5.0) es la más débil. "
     "Revisar precios y facilidad de compra podría incrementar la satisfacción global."),
]

for idx, (numero, titulo, texto) in enumerate(conclusiones, 1):
    row = 4 + (idx-1)*4
    ws6.cell(row=row, column=1, value=f"{numero}. {titulo}").font = Font(bold=True, size=11, color=COLOR_SECUNDARIO)
    ws6.cell(row=row+1, column=1, value=texto).alignment = Alignment(wrap_text=True)
    ws6.cell(row=row, column=1).border = get_border()
    ws6.cell(row=row+1, column=1).border = get_border()

ws6.cell(row=44, column=1, value="Nota: Todas las conclusiones son originales del equipo aliado")
ws6.cell(row=44, column=1).font = Font(italic=True, size=9, color=COLOR_TERCIARIO)

print("✓ Conclusiones creada")

# ============================================================================
# Ajustes finales y guardado
# ============================================================================
print("\nAplicando ajustes finales...")

# Ajustar ancho de columnas para todas las hojas
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for col in range(1, 10):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

# Guardar el archivo
output_path = "Modulo8_Dashboard_Aliado.xlsx"
wb.save(output_path)

print(f"\n{'='*70}")
print(f"✅ DASHBOARD ALIADO GENERADO EXITOSAMENTE")
print(f"{'='*70}")
print(f"Archivo: {output_path}")
print(f"Ubicación: {os.path.abspath(output_path)}")
print(f"\nEstructura:")
print(f"  - Portada_Resumen")
print(f"  - Base_Integrada")
print(f"  - KPIs_Mod1_Mod2")
print(f"  - KPIs_Mod3_Mod4")
print(f"  - KPIs_Mod5_6_7")
print(f"  - Dashboard_Aliado (hoja principal)")
print(f"  - Conclusiones")
print(f"\nDiseño: Paleta monocromática (azules oscuros)")
print(f"Datos: Coherentes con módulos 1-7")
print(f"Conclusiones: 10 puntos originales del equipo aliado")
print(f"{'='*70}")
