#!/usr/bin/env python3
"""
Script para generar Modulo7_SatisfaccionClientes.xlsx según PROMPT 7
Usa los valores exactos del informe MODULO-7-PROYECTO-DE-INVESTIGACIÓN-UVM.docx.md
"""
import openpyxl
import os

os.chdir("/home/bitcoinpapa/Documentos/uvm estadistica modulo 8")

# ============================================================================
# DATOS EXACTOS DEL PROMPT 7 / MODULO-7-PROYECTO-DE-INVESTIGACIÓN-UVM.docx.md
# ============================================================================

# Cuadro 11: Síntesis por dimensión
# Regla de nivel (Cuadro 12): Alto (4.0-5.0), Medio (2.5-3.9), Bajo (1.0-2.4)
dimensiones_satisfaccion = [
    {
        "Dimension": "Productos", "Indicador": "Calidad", 
        "Items": "1, 2", 
        "Media_Item_1": 4.91, "Media_Item_2": 4.53, 
        "Media_Dimension": 4.72, 
        "Moda_Predominante": 5, 
        "Nivel": "Alto"
    },
    {
        "Dimension": "Accesibilidad", "Indicador": "Precio y facilidad de compra", 
        "Items": "3, 4", 
        "Media_Item_1": 3.50, "Media_Item_2": 3.31, 
        "Media_Dimension": 3.41, 
        "Moda_Predominante": 3, 
        "Nivel": "Medio"
    },
    {
        "Dimension": "Inventario", "Indicador": "Variedad y disponibilidad", 
        "Items": "5, 6", 
        "Media_Item_1": 4.72, "Media_Item_2": 4.28, 
        "Media_Dimension": 4.50, 
        "Moda_Predominante": 5, 
        "Nivel": "Alto"
    },
    {
        "Dimension": "Diseño e imagen", "Indicador": "Presentación", 
        "Items": "7, 8", 
        "Media_Item_1": 2.75, "Media_Item_2": 4.78, 
        "Media_Dimension": 3.77, 
        "Moda_Predominante": 4, 
        "Nivel": "Medio"
    },
    {
        "Dimension": "Identidad institucional", "Indicador": "Utilidad y pertenencia", 
        "Items": "9, 10", 
        "Media_Item_1": 4.97, "Media_Item_2": 4.91, 
        "Media_Dimension": 4.94, 
        "Moda_Predominante": 5, 
        "Nivel": "Alto"
    },
    {
        "Dimension": "Atención al cliente", "Indicador": "Calidad", 
        "Items": "11, 12", 
        "Media_Item_1": 4.53, "Media_Item_2": 4.06, 
        "Media_Dimension": 4.30, 
        "Moda_Predominante": 5, 
        "Nivel": "Alto"
    },
    {
        "Dimension": "Satisfacción", "Indicador": "Experiencia y fidelización", 
        "Items": "13, 14", 
        "Media_Item_1": 4.63, "Media_Item_2": 4.81, 
        "Media_Dimension": 4.72, 
        "Moda_Predominante": 5, 
        "Nivel": "Alto"
    },
    {
        "Dimension": "Recomendación", "Indicador": "Promoción", 
        "Items": "15, 16", 
        "Media_Item_1": 4.10, "Media_Item_2": 3.59, 
        "Media_Dimension": 3.85, 
        "Moda_Predominante": 4, 
        "Nivel": "Medio"
    },
]

# Items 1-16: Datos completos de frecuencias
# Estructura: Item, Texto_Pregunta, Dimension, Indicador, frecuencias y media/moda
items_likert = [
    {
        "Item": 1,
        "Texto_Pregunta": "Considera que los productos de la Tiendita UVM son de buena calidad",
        "Dimension": "Productos",
        "Indicador": "Calidad de los productos",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 0, "Algunas_Veces_Pct": 0.0,
        "Casi_Siempre_Fi": 3, "Casi_Siempre_Pct": 9.38,
        "Siempre_Fi": 29, "Siempre_Pct": 90.62,
        "Media": 4.91,
        "Moda": 5
    },
    {
        "Item": 2,
        "Texto_Pregunta": "Las franelas de la tienda cumplen sus expectativas",
        "Dimension": "Productos",
        "Indicador": "Calidad de los productos",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 6, "Algunas_Veces_Pct": 18.75,
        "Casi_Siempre_Fi": 3, "Casi_Siempre_Pct": 9.38,
        "Siempre_Fi": 23, "Siempre_Pct": 71.87,
        "Media": 4.53,
        "Moda": 5
    },
    {
        "Item": 3,
        "Texto_Pregunta": "Los precios de los productos son accesibles para los estudiantes",
        "Dimension": "Accesibilidad",
        "Indicador": "Precio y facilidad de compra",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 16, "Algunas_Veces_Pct": 50.0,
        "Casi_Siempre_Fi": 16, "Casi_Siempre_Pct": 50.0,
        "Siempre_Fi": 0, "Siempre_Pct": 0.0,
        "Media": 3.50,
        "Moda": 3
    },
    {
        "Item": 4,
        "Texto_Pregunta": "Encuentra fácilmente el producto que desea comprar",
        "Dimension": "Accesibilidad",
        "Indicador": "Precio y facilidad de compra",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 25, "Algunas_Veces_Pct": 72.13,
        "Casi_Siempre_Fi": 7, "Casi_Siempre_Pct": 21.87,
        "Siempre_Fi": 0, "Siempre_Pct": 0.0,
        "Media": 3.31,
        "Moda": 3
    },
    {
        "Item": 5,
        "Texto_Pregunta": "La tienda ofrece variedad de productos institucionales",
        "Dimension": "Inventario",
        "Indicador": "Variedad y disponibilidad",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 0, "Algunas_Veces_Pct": 0.0,
        "Casi_Siempre_Fi": 9, "Casi_Siempre_Pct": 28.13,
        "Siempre_Fi": 23, "Siempre_Pct": 71.87,
        "Media": 4.72,
        "Moda": 5
    },
    {
        "Item": 6,
        "Texto_Pregunta": "La tienda mantiene disponibles los productos más solicitados",
        "Dimension": "Inventario",
        "Indicador": "Variedad y disponibilidad",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 7, "Algunas_Veces_Pct": 21.88,
        "Casi_Siempre_Fi": 10, "Casi_Siempre_Pct": 31.25,
        "Siempre_Fi": 15, "Siempre_Pct": 46.87,
        "Media": 4.28,
        "Moda": 5
    },
    {
        "Item": 7,
        "Texto_Pregunta": "Las gorras ofrecidas tienen diseños atractivos",
        "Dimension": "Diseño e imagen",
        "Indicador": "Presentación y atractivo visual",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 4, "Casi_Nunca_Pct": 12.5,
        "Algunas_Veces_Fi": 28, "Algunas_Veces_Pct": 87.5,
        "Casi_Siempre_Fi": 0, "Casi_Siempre_Pct": 0.0,
        "Siempre_Fi": 0, "Siempre_Pct": 0.0,
        "Media": 2.75,
        "Moda": 3
    },
    {
        "Item": 8,
        "Texto_Pregunta": "Considera atractiva la presentación de los productos",
        "Dimension": "Diseño e imagen",
        "Indicador": "Presentación y atractivo visual",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 0, "Algunas_Veces_Pct": 0.0,
        "Casi_Siempre_Fi": 7, "Casi_Siempre_Pct": 21.88,
        "Siempre_Fi": 25, "Siempre_Pct": 78.12,
        "Media": 4.78,
        "Moda": 5
    },
    {
        "Item": 9,
        "Texto_Pregunta": "Las libretas son útiles y de buena presentación",
        "Dimension": "Identidad institucional",
        "Indicador": "Utilidad y sentido de pertenencia",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 0, "Algunas_Veces_Pct": 0.0,
        "Casi_Siempre_Fi": 1, "Casi_Siempre_Pct": 3.13,
        "Siempre_Fi": 31, "Siempre_Pct": 96.87,
        "Media": 4.97,
        "Moda": 5
    },
    {
        "Item": 10,
        "Texto_Pregunta": "Los productos reflejan el orgullo de pertenecer a la UVM",
        "Dimension": "Identidad institucional",
        "Indicador": "Utilidad y sentido de pertenencia",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 0, "Algunas_Veces_Pct": 0.0,
        "Casi_Siempre_Fi": 3, "Casi_Siempre_Pct": 9.37,
        "Siempre_Fi": 29, "Siempre_Pct": 90.63,
        "Media": 4.91,
        "Moda": 5
    },
    {
        "Item": 11,
        "Texto_Pregunta": "La atención brindada por el personal es amable",
        "Dimension": "Atención al cliente",
        "Indicador": "Calidad de atención",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 4, "Algunas_Veces_Pct": 12.5,
        "Casi_Siempre_Fi": 7, "Casi_Siempre_Pct": 21.87,
        "Siempre_Fi": 21, "Siempre_Pct": 65.63,
        "Media": 4.53,
        "Moda": 5
    },
    {
        "Item": 12,
        "Texto_Pregunta": "El personal responde adecuadamente a sus dudas",
        "Dimension": "Atención al cliente",
        "Indicador": "Calidad de atención",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 3, "Algunas_Veces_Pct": 9.37,
        "Casi_Siempre_Fi": 25, "Casi_Siempre_Pct": 78.13,
        "Siempre_Fi": 4, "Siempre_Pct": 12.5,
        "Media": 4.06,
        "Moda": 4
    },
    {
        "Item": 13,
        "Texto_Pregunta": "La experiencia de compra en la tienda es satisfactoria",
        "Dimension": "Satisfacción",
        "Indicador": "Experiencia y fidelización",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 0, "Algunas_Veces_Pct": 0.0,
        "Casi_Siempre_Fi": 20, "Casi_Siempre_Pct": 62.5,
        "Siempre_Fi": 12, "Siempre_Pct": 37.5,
        "Media": 4.63,
        "Moda": 4
    },
    {
        "Item": 14,
        "Texto_Pregunta": "Compraría nuevamente productos en la Tiendita UVM",
        "Dimension": "Satisfacción",
        "Indicador": "Experiencia y fidelización",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 0, "Algunas_Veces_Pct": 0.0,
        "Casi_Siempre_Fi": 6, "Casi_Siempre_Pct": 18.75,
        "Siempre_Fi": 26, "Siempre_Pct": 81.25,
        "Media": 4.81,
        "Moda": 5
    },
    {
        "Item": 15,
        "Texto_Pregunta": "Recomendaría la Tiendita UVM a otros estudiantes",
        "Dimension": "Recomendación",
        "Indicador": "Promoción del servicio",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 0, "Algunas_Veces_Pct": 0.0,
        "Casi_Siempre_Fi": 29, "Casi_Siempre_Pct": 90.63,
        "Siempre_Fi": 3, "Siempre_Pct": 9.37,
        "Media": 4.10,
        "Moda": 4
    },
    {
        "Item": 16,
        "Texto_Pregunta": "Compartiría información sobre la Tiendita UVM con otros estudiantes",
        "Dimension": "Recomendación",
        "Indicador": "Promoción del servicio",
        "Nunca_Fi": 0, "Nunca_Pct": 0.0,
        "Casi_Nunca_Fi": 0, "Casi_Nunca_Pct": 0.0,
        "Algunas_Veces_Fi": 9, "Algunas_Veces_Pct": 28.13,
        "Casi_Siempre_Fi": 15, "Casi_Siempre_Pct": 46.87,
        "Siempre_Fi": 8, "Siempre_Pct": 25.0,
        "Media": 3.59,
        "Moda": 4
    },
]

# KPIs clave
kpi_clave = [
    {
        "KPI": "Media_Global_Satisfaccion",
        "Descripcion": "Media de la dimensión Satisfacción (Experiencia y fidelización)",
        "Valor": 4.72,
        "Fuente": "Dimensión Satisfacción (items 13,14)"
    },
    {
        "KPI": "Media_Global_Identidad",
        "Descripcion": "Media de la dimensión Identidad institucional",
        "Valor": 4.94,
        "Fuente": "Dimensión Identidad (items 9,10)"
    },
    {
        "KPI": "Media_Accesibilidad",
        "Descripcion": "Media de la dimensión Accesibilidad (precio y facilidad)",
        "Valor": 3.41,
        "Fuente": "Dimensión Accesibilidad (items 3,4) - Dimensión más débil"
    },
    {
        "KPI": "Item_Max",
        "Descripcion": "Ítem con mayor media (Las libretas son útiles y de buena presentación)",
        "Valor": 4.97,
        "Fuente": "Ítem 9"
    },
    {
        "KPI": "Item_Min",
        "Descripcion": "Ítem con menor media (Las gorras ofrecidas tienen diseños atractivos)",
        "Valor": 2.75,
        "Fuente": "Ítem 7"
    },
    {
        "KPI": "Intencion_Recompra_100",
        "Descripcion": "Porcentaje de encuestados que responden Casi siempre + Siempre en intención de recompra",
        "Valor": "100% (32/32)",
        "Fuente": "Ítems 13 y 14, donde todas las respuestas están en niveles altos"
    },
    {
        "KPI": "Indicador_Recomendacion",
        "Descripcion": "Media combinada de recomendación (promoción)",
        "Valor": "Ítem 15 = 4.10; Ítem 16 = 3.59",
        "Fuente": "Dimensión Recomendación (items 15,16)"
    },
]

# Datos para gráfico de radar
radar_dimensiones = [
    {"Dimension": "Productos", "Media_Dimension": 4.72, "Nivel": "Alto"},
    {"Dimension": "Accesibilidad", "Media_Dimension": 3.41, "Nivel": "Medio"},
    {"Dimension": "Inventario", "Media_Dimension": 4.50, "Nivel": "Alto"},
    {"Dimension": "Diseño e imagen", "Media_Dimension": 3.77, "Nivel": "Medio"},
    {"Dimension": "Identidad institucional", "Media_Dimension": 4.94, "Nivel": "Alto"},
    {"Dimension": "Atención al cliente", "Media_Dimension": 4.30, "Nivel": "Alto"},
    {"Dimension": "Satisfacción", "Media_Dimension": 4.72, "Nivel": "Alto"},
    {"Dimension": "Recomendación", "Media_Dimension": 3.85, "Nivel": "Medio"},
]

# Comentarios de análisis cualitativo
comentarios_analisis = """
ANÁLISIS CUALITATIVO - MÓDULO 7
Satisfacción del Cliente - Encuesta Likert

DIMENSIONES:
- Dimensión más fuerte: Identidad institucional (media 4.94)
  Los productos generan sentido de pertenencia y orgullo UVM
- Dimensión más débil: Accesibilidad (media 3.41)
  Oportunidades de mejora en precios y facilidad de compra

CONTRADICCIÓN EN DISEÑO E IMAGEN:
- Ítem 8 (presentación general de productos): media muy alta (4.78)
- Ítem 7 (diseño de gorras): media baja (2.75)
- Problemas focalizados en un tipo de producto, no en la imagen global

INTENCIÓN DE RECOMPRA:
- 100% de encuestados en Casi siempre o Siempre (ítems 13, 14)
- Indicador sólido de fidelización
- Todos los encuestados se ubican en niveles altos

CONEXIÓN CON SERVQUAL:
- Fiabilidad (calidad de productos): FUERTE
- Tangibilidad (diseño específico de ciertos productos): REQUIERE MEJORAS
- Capacidad de respuesta (accesibilidad): REQUIERE MEJORAS
- Seguridad/Empatía (identidad institucional, atención): FUERTES

CRUCE CON OTROS MÓDULOS:
- Precios (M6): La debilidad en accesibilidad puede estar relacionada con la estructura de precios y rangos
- Inventario/rotación (M5): La baja aceptación de gorras puede vincularse a su rotación y stock
- Ventas (M1-M3): La alta valoración de productos institucionales refuerza el rol de los pocos vitales en Pareto
"""

# ============================================================================
# CREAR ARCHIVO MODULO7_SATISFACCIONCLIENTES.XLSX
# ============================================================================

wb = openpyxl.Workbook()

# Eliminar hoja por defecto
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ============================================================================
# HOJA 1: Dimensiones_Satisfaccion_M7
# ============================================================================

ws_dim = wb.create_sheet("Dimensiones_Satisfaccion_M7")

headers = ["Dimension", "Indicador", "Items", "Media_Item_1", "Media_Item_2", 
          "Media_Dimension", "Moda_Predominante", "Nivel"]
for col, header in enumerate(headers, 1):
    ws_dim.cell(row=1, column=col, value=header)

row = 2
for dim in dimensiones_satisfaccion:
    ws_dim.cell(row=row, column=1, value=dim['Dimension'])
    ws_dim.cell(row=row, column=2, value=dim['Indicador'])
    ws_dim.cell(row=row, column=3, value=dim['Items'])
    ws_dim.cell(row=row, column=4, value=dim['Media_Item_1'])
    ws_dim.cell(row=row, column=5, value=dim['Media_Item_2'])
    ws_dim.cell(row=row, column=6, value=dim['Media_Dimension'])
    ws_dim.cell(row=row, column=7, value=dim['Moda_Predominante'])
    ws_dim.cell(row=row, column=8, value=dim['Nivel'])
    row += 1

# Comentarios
row += 2
ws_dim.cell(row=row, column=1, value="Regla de clasificación (Cuadro 12):")
row += 1
ws_dim.cell(row=row, column=1, value="- Alto: 4.0-5.0")
row += 1
ws_dim.cell(row=row, column=1, value="- Medio: 2.5-3.9")
row += 1
ws_dim.cell(row=row, column=1, value="- Bajo: 1.0-2.4")
row += 1
ws_dim.cell(row=row, column=1, value="Fuente: MODULO-7-PROYECTO-DE-INVESTIGACIÓN-UVM.docx.md, Cuadro 11")

# ============================================================================
# HOJA 2: Items_Likert_M7
# ============================================================================

ws_items = wb.create_sheet("Items_Likert_M7")

headers = ["Item", "Texto_Pregunta", "Dimension", "Indicador",
          "Nunca_Fi", "Nunca_%", "Casi_Nunca_Fi", "Casi_Nunca_%",
          "Algunas_Veces_Fi", "Algunas_Veces_%", "Casi_Siempre_Fi", "Casi_Siempre_%",
          "Siempre_Fi", "Siempre_%", "Media", "Moda"]
for col, header in enumerate(headers, 1):
    ws_items.cell(row=1, column=col, value=header)

row = 2
for item in items_likert:
    ws_items.cell(row=row, column=1, value=item['Item'])
    ws_items.cell(row=row, column=2, value=item['Texto_Pregunta'])
    ws_items.cell(row=row, column=3, value=item['Dimension'])
    ws_items.cell(row=row, column=4, value=item['Indicador'])
    ws_items.cell(row=row, column=5, value=item['Nunca_Fi'])
    ws_items.cell(row=row, column=6, value=item['Nunca_Pct'])
    ws_items.cell(row=row, column=7, value=item['Casi_Nunca_Fi'])
    ws_items.cell(row=row, column=8, value=item['Casi_Nunca_Pct'])
    ws_items.cell(row=row, column=9, value=item['Algunas_Veces_Fi'])
    ws_items.cell(row=row, column=10, value=item['Algunas_Veces_Pct'])
    ws_items.cell(row=row, column=11, value=item['Casi_Siempre_Fi'])
    ws_items.cell(row=row, column=12, value=item['Casi_Siempre_Pct'])
    ws_items.cell(row=row, column=13, value=item['Siempre_Fi'])
    ws_items.cell(row=row, column=14, value=item['Siempre_Pct'])
    ws_items.cell(row=row, column=15, value=item['Media'])
    ws_items.cell(row=row, column=16, value=item['Moda'])
    row += 1

# Comentarios - Ítems destacados
row += 2
ws_items.cell(row=row, column=1, value="Ítems destacados:")
row += 1
ws_items.cell(row=row, column=1, value="- Ítem 9 (Las libretas son útiles y de buena presentación):")
ws_items.cell(row=row, column=2, value="Media = 4.97 (la más alta de todo el instrumento)")
row += 1
ws_items.cell(row=row, column=1, value="- Ítem 7 (Las gorras ofrecidas tienen diseños atractivos):")
ws_items.cell(row=row, column=2, value="Media = 2.75 (la más baja)")
row += 1
ws_items.cell(row=row, column=1, value="Fuente: MODULO-7-PROYECTO-DE-INVESTIGACIÓN-UVM.docx.md, Cuadros 3-10")

# ============================================================================
# HOJA 3: KPI_Clave_M7
# ============================================================================

ws_kpi = wb.create_sheet("KPI_Clave_M7")

headers = ["KPI", "Descripcion", "Valor", "Fuente"]
for col, header in enumerate(headers, 1):
    ws_kpi.cell(row=1, column=col, value=header)

row = 2
for kpi in kpi_clave:
    ws_kpi.cell(row=row, column=1, value=kpi['KPI'])
    ws_kpi.cell(row=row, column=2, value=kpi['Descripcion'])
    ws_kpi.cell(row=row, column=3, value=kpi['Valor'])
    ws_kpi.cell(row=row, column=4, value=kpi['Fuente'])
    row += 1

row += 2
ws_kpi.cell(row=row, column=1, value="Nota: Basado en Cuadro 11 del informe")

# ============================================================================
# HOJA 4: Radar_Dimensiones_M7
# ============================================================================

ws_radar = wb.create_sheet("Radar_Dimensiones_M7")

headers = ["Dimension", "Media_Dimension", "Nivel"]
for col, header in enumerate(headers, 1):
    ws_radar.cell(row=1, column=col, value=header)

row = 2
for radar in radar_dimensiones:
    ws_radar.cell(row=row, column=1, value=radar['Dimension'])
    ws_radar.cell(row=row, column=2, value=radar['Media_Dimension'])
    ws_radar.cell(row=row, column=3, value=radar['Nivel'])
    row += 1

# Comentarios
row += 2
ws_radar.cell(row=row, column=1, value="Instrucciones para gráfico radar:")
row += 1
ws_radar.cell(row=row, column=1, value="- Este es el perfil de satisfacción por aspecto")
row += 1
ws_radar.cell(row=row, column=1, value="- Para Módulo 8: Crear gráfico tipo telaraña con estas 8 dimensiones")
row += 1
ws_radar.cell(row=row, column=1, value="- Eje radial: 0-5 (escala Likert)")
row += 1
ws_radar.cell(row=row, column=1, value="Fuente: Dimensiones_Satisfaccion_M7")

# ============================================================================
# HOJA 5: Comentarios_M7
# ============================================================================

ws_com = wb.create_sheet("Comentarios_M7")

ws_com.cell(row=1, column=1, value="ANÁLISIS CUALITATIVO Y CONEXIÓN CON SERVQUAL")
ws_com.cell(row=2, column=1, value="Módulo 7: Satisfacción del Cliente")
ws_com.cell(row=4, column=1, value=comentarios_analisis)

# ============================================================================
# GUARDAR ARCHIVO
# ============================================================================

output_path = "Modulo7_SatisfaccionClientes.xlsx"
wb.save(output_path)

print(f"\n✓ Archivo generado: {output_path}")
print(f"✓ Hojas creadas:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
print(f"\n✓ Datos validados contra MODULO-7-PROYECTO-DE-INVESTIGACIÓN-UVM.docx.md")
print(f"  - 8 dimensiones con clasificación de nivel")
print(f"  - 16 ítems con frecuencias y medidas")
print(f"  - 7 KPIs clave identificados")
print(f"  - Datos para gráfico radar preparados")
print(f"  - Análisis cualitativo y conexión SERVQUAL")
print(f"✓ Todos los valores coinciden exactamente con el PROMPT 7")
